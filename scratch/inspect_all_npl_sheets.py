import openpyxl
import os

files = [
    './PIPELINE PROPOSTAS - PIETRA.xlsx',
    './CASOS_NPL_FECHADOS_CONSOLIDADO_INTEGRADO.xlsx',
    'C:/Users/ArthurFeltrinDeco/OneDrive - Lepta/Arquivos de Chat do Microsoft Teams/CASOS_NPL_FECHADOS_CONSOLIDADO_INTEGRADO.xlsx'
]

for f in files:
    if not os.path.exists(f):
        continue
    print(f"\n==================== ARQUIVO: {f} ====================")
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} ---")
        rows = list(ws.iter_rows(max_row=12, values_only=True))
        for idx, r in enumerate(rows):
            non_empty = [f"[Col {c_idx+1}] {str(val)[:25]}" for c_idx, val in enumerate(r) if val is not None and str(val).strip() != '']
            if non_empty:
                print(f"Row {idx+1}: {' | '.join(non_empty)}")
